import os
import asyncio
import re
import sqlite3
import psycopg2
import openpyxl
from dotenv import load_dotenv
from playwright.async_api import async_playwright

# Load environmental variables
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
load_dotenv(os.path.join(BASE_DIR, "../backend/.env"))

POSTGRES_URL = os.environ.get(
    "DATABASE_URL", 
    "postgresql://penner_admin:penner_secret_password_2026@localhost:5432/penner_governance_db"
)
SQLITE_PATH = "/Users/thejoshuapenner/My Drive/Penner Strategy/sao-scraper/municipal_intent.db"

def get_pg_conn():
    try:
        return psycopg2.connect(POSTGRES_URL)
    except Exception as e:
        print(f"Postgres connection failed: {e}")
        return None

def get_sqlite_conn():
    try:
        conn = sqlite3.connect(SQLITE_PATH)
        conn.row_factory = sqlite3.Row
        conn.execute("PRAGMA journal_mode=DELETE;")
        conn.execute("PRAGMA busy_timeout = 30000;")
        return conn
    except Exception as e:
        print(f"SQLite connection failed: {e}")
        return None

def load_jurisdictions(conn, is_pg: bool) -> dict:
    """Loads standardized jurisdictions mapping to clean names and entity types."""
    cur = conn.cursor()
    jurisdictions = {}
    try:
        cur.execute("SELECT name, entity_type FROM jurisdictions")
        rows = cur.fetchall()
        for r in rows:
            name = r[0] if is_pg else r['name']
            etype = r[1] if is_pg else r['entity_type']
            
            clean_name = name.strip()
            if clean_name.startswith('"') and clean_name.endswith('"'):
                clean_name = clean_name[1:-1].strip()
                
            jurisdictions[clean_name.upper()] = {
                'std_name': clean_name,
                'entity_type': etype
            }
        return jurisdictions
    except Exception as e:
        print(f"Failed loading jurisdictions: {e}")
        return {}
    finally:
        cur.close()

PORT_ALIASES = {
    "PORT OF COLUMBIA COUNTY": "Port of Columbia",
    "PORT OF GARFIELD COUNTY": "Port of Garfield",
    "PORT OF QUINCY": "Port of Grant County No. 1 (Quincy)",
    "PORT OF MOSES LAKE": "Port of Grant County No. 10 (Moses Lake)",
    "PORT OF ROYAL SLOPE": "Port of Grant County No. 2 (Royal Slope)",
    "PORT OF MATTAWA": "Port of Grant County No. 3 (Mattawa)",
    "PORT OF COULEE CITY": "Port of Grant County No. 4 (Coulee City)",
    "PORT OF HARTLINE": "Port of Grant County No. 5 (Hartline)",
    "PORT OF WILSON CREEK": "Port of Grant County No. 6 (Wilson Creek)",
    "PORT OF GRAND COULEE": "Port of Grant County No. 7 (Grand Coulee)",
    "PORT OF WARDEN": "Port of Grant County No. 8 (Warden)",
    "PORT OF EPHRATA": "Port of Grant County No. 9 (Ephrata)",
    "KLICKITAT COUNTY PORT DISTRICT NO. 1": "Port of Klickitat",
    "WAHKIAKUM COUNTY PORT DISTRICT NO. 1": "Port of Wahkiakum County No. 1",
    "WAHKIAKUM COUNTY PORT DISTRICT NO. 2": "Port of Wahkiakum County No. 2",
    "PEND OREILLE VALLEY RAILROAD": "Port of Pend Oreille",
}

def match_jurisdiction(common_name, jurisdictions) -> tuple:
    """Matches SAO common name to our standardized jurisdiction list strictly."""
    if not common_name:
        return None, None
        
    name_up = common_name.strip().upper()
    
    # 1. Check for specific Port Aliases first
    if name_up in PORT_ALIASES:
        std_name = PORT_ALIASES[name_up]
        std_name_up = std_name.upper()
        if std_name_up in jurisdictions:
            return jurisdictions[std_name_up]['std_name'], 'port'
        return std_name, 'port'
        
    # 2. Check for Town of Saint John -> St. John city translation
    if name_up == "TOWN OF SAINT JOHN" or name_up == "SAINT JOHN":
        if "ST. JOHN" in jurisdictions:
            return jurisdictions["ST. JOHN"]['std_name'], 'city'
            
    # 3. Direct matches
    if name_up in jurisdictions:
        return jurisdictions[name_up]['std_name'], jurisdictions[name_up]['entity_type']
        
    # 4. Clean up Cities (strip "City of " or "Town of " from FIT name)
    if name_up.startswith("CITY OF "):
        base = name_up[8:].strip()
        if base in jurisdictions and jurisdictions[base]['entity_type'] == 'city':
            return jurisdictions[base]['std_name'], 'city'
    elif name_up.startswith("TOWN OF "):
        base = name_up[8:].strip()
        if base in jurisdictions and jurisdictions[base]['entity_type'] == 'city':
            return jurisdictions[base]['std_name'], 'city'
            
    # 5. Clean up Counties
    if name_up.endswith(" COUNTY"):
        if name_up in jurisdictions and jurisdictions[name_up]['entity_type'] == 'county':
            return jurisdictions[name_up]['std_name'], 'county'
            
    # 6. Clean up Ports
    if name_up.startswith("PORT OF "):
        if name_up in jurisdictions and jurisdictions[name_up]['entity_type'] == 'port':
            return jurisdictions[name_up]['std_name'], 'port'
            
    return None, None

async def download_sao_extract(year) -> str:
    """Uses Playwright to download the full single-year FIT data extract from the SAO website."""
    url = "https://portal.sao.wa.gov/FIT/"
    scratch_dir = "/Users/thejoshuapenner/.gemini/antigravity/brain/a79b3e23-62f8-4a39-a5f2-fafeec1789e0/scratch"
    os.makedirs(scratch_dir, exist_ok=True)
    temp_file = os.path.join(scratch_dir, f"fit_{year}.xlsx")
    
    print(f"  [PLAYWRIGHT] Downloading FIT data extract for Year {year}...")
    
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        page = await browser.new_page()
        
        try:
            await page.goto(url, timeout=60000)
            await page.wait_for_timeout(3000)
            
            await page.click("text=Data Extracts")
            await page.wait_for_timeout(3000)
            
            await page.click("text=Full Extracts")
            await page.wait_for_timeout(4000)
            
            card_text = f"{year} Data"
            print(f"  [PLAYWRIGHT] Clicking card: '{card_text}'...")
            
            async with page.expect_download(timeout=90000) as download_info:
                await page.click(f"h5:has-text('{card_text}')")
                
            download = await download_info.value
            await download.save_as(temp_file)
            print(f"  [PLAYWRIGHT] Successfully downloaded {download.suggested_filename} to {temp_file}")
            print(f"  [PLAYWRIGHT] Size: {os.path.getsize(temp_file):,} bytes")
            return temp_file
        except Exception as e:
            print(f"  [PLAYWRIGHT ERROR] Download failed for Year {year}: {e}")
            return None
        finally:
            await browser.close()

def parse_and_ingest_fit_file(file_path, year):
    """Parses the downloaded Excel file and ingests the General Fund budgets into the databases."""
    print(f"  [PARSER] Loading excel workbook: {file_path} (this may take a few seconds)...")
    wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    
    # 1. Load Ref_Governments
    print("  [PARSER] Loading Ref_Governments...")
    gov_map = {}
    if 'Ref_Governments' in wb.sheetnames:
        sheet = wb['Ref_Governments']
        headers = []
        for row in sheet.iter_rows(values_only=True):
            if not headers:
                headers = row
                continue
            r_dict = dict(zip(headers, row))
            mcag = r_dict.get('MCAG')
            name = r_dict.get('CommonName', r_dict.get('LegalName'))
            if mcag and name:
                gov_map[str(mcag)] = name
                
    # 2. Load Ref_AccountDesc
    print("  [PARSER] Loading Ref_AccountDesc...")
    acct_map = {}
    if 'Ref_AccountDesc' in wb.sheetnames:
        sheet = wb['Ref_AccountDesc']
        headers = []
        for row in sheet.iter_rows(values_only=True):
            if not headers:
                headers = row
                continue
            r_dict = dict(zip(headers, row))
            acct_id = r_dict.get('Id')
            logical = r_dict.get('LogicalAccount')
            name = r_dict.get('Name')
            if acct_id and logical and name:
                acct_map[int(acct_id)] = {
                    'code': str(logical),
                    'name': name
                }
                
    # Load jurisdictions to filter/match target universe (brief DB connection)
    sqlite_conn = get_sqlite_conn()
    pg_conn = get_pg_conn()
    conn_to_use = pg_conn if pg_conn else sqlite_conn
    is_pg = True if pg_conn else False
    if not conn_to_use:
        print("  [ERROR] Database connections not available.")
        wb.close()
        return
    jurisdictions = load_jurisdictions(conn_to_use, is_pg)
    if sqlite_conn:
        sqlite_conn.close()
    if pg_conn:
        pg_conn.close()
        
    sheet_name = f'Schedule1_{year}'
    if sheet_name not in wb.sheetnames:
        sheet_name = wb.sheetnames[1]
        
    print(f"  [PARSER] Processing financial sheet: {sheet_name}...")
    sheet = wb[sheet_name]
    
    parent_budgets = {}
    headers = []
    
    for row in sheet.iter_rows(values_only=True):
        if not headers:
            headers = row
            continue
            
        r_dict = dict(zip(headers, row))
        
        # Filter for General Fund
        fund_number = str(r_dict.get('FundNumber', '')).strip()
        fund_name = str(r_dict.get('FundName', '')).strip()
        
        if fund_number != '001' and 'GENERAL' not in fund_name.upper() and 'CURRENT EXPENSE' not in fund_name.upper():
            continue
            
        mcag = str(r_dict.get('MCAG', ''))
        amount = float(r_dict.get('Amount', 0.0))
        acct_id = r_dict.get('BARSAccountId')
        
        if not mcag or amount <= 0 or acct_id is None:
            continue
            
        gov_raw_name = gov_map.get(mcag)
        acct_info = acct_map.get(int(acct_id))
        
        if not gov_raw_name or not acct_info:
            continue
            
        std_name, etype = match_jurisdiction(gov_raw_name, jurisdictions)
        if not std_name:
            continue
            
        code = acct_info['code']
        label = acct_info['name']
        
        if mcag not in parent_budgets:
            parent_budgets[mcag] = {
                'std_name': std_name,
                'entity_type': etype,
                'total_revenue': 0.0,
                'total_expenditures': 0.0,
                'fund_balance_beginning': 0.0,
                'fund_balance_ending': 0.0,
                'items': []
            }
            
        p = parent_budgets[mcag]
        
        if code.startswith('3'):
            if code.startswith('308'):
                p['fund_balance_beginning'] += amount
            else:
                p['total_revenue'] += amount
                p['items'].append(('revenue', label, amount, f"General Fund Revenue: {label}"))
        elif code.startswith('5'):
            if code.startswith('508'):
                p['fund_balance_ending'] += amount
            else:
                p['total_expenditures'] += amount
                p['items'].append(('expenditure', label, amount, f"General Fund Expenditure: {label}"))
                
    wb.close()
    
    print(f"  [PARSER] Finished file analysis. Aligned {len(parent_budgets)} budgets to insert.")
    
    # Re-open DB connections briefly for one fast transaction
    sqlite_conn = get_sqlite_conn()
    pg_conn = get_pg_conn()
    
    sqlite_cur = sqlite_conn.cursor() if sqlite_conn else None
    pg_cur = pg_conn.cursor() if pg_conn else None
    
    saved_budget_count = 0
    saved_item_count = 0
    source_url = f"https://portal.sao.wa.gov/FIT/data-extracts/full"
    
    # 3. Process database inserts in one fast transaction block
    if sqlite_conn and sqlite_cur:
        print("  [DB] Writing to SQLite...")
        try:
            # We do all SELECT and INSERT statements inside a fast transaction
            sqlite_cur.execute("BEGIN TRANSACTION")
            
            sqlite_batch_items = []
            
            for mcag, data in parent_budgets.items():
                std_name = data['std_name']
                etype = data['entity_type']
                tot_rev = data['total_revenue']
                tot_exp = data['total_expenditures']
                beg_bal = data['fund_balance_beginning']
                end_bal = data['fund_balance_ending']
                
                sqlite_budget_id = None
                
                sqlite_cur.execute(
                    "SELECT id FROM budgets WHERE jurisdiction_name = ? AND entity_type = ? AND fiscal_year = ? LIMIT 1",
                    (std_name, etype, year)
                )
                row = sqlite_cur.fetchone()
                if row:
                    sqlite_budget_id = row[0]
                    sqlite_cur.execute(
                        """
                        UPDATE budgets 
                        SET total_revenue = ?, total_expenditures = ?, fund_balance_beginning = ?, fund_balance_ending = ?, source_url = ?
                        WHERE id = ?
                        """,
                        (tot_rev, tot_exp, beg_bal, end_bal, source_url, sqlite_budget_id)
                    )
                else:
                    sqlite_cur.execute(
                        """
                        INSERT INTO budgets (jurisdiction_name, entity_type, fiscal_year, total_revenue, total_expenditures, fund_balance_beginning, fund_balance_ending, source_url)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                        """,
                        (std_name, etype, year, tot_rev, tot_exp, beg_bal, end_bal, source_url)
                    )
                    sqlite_budget_id = sqlite_cur.lastrowid
                    saved_budget_count += 1
                    
                if sqlite_budget_id and data['items']:
                    # Clean up existing general fund items for this budget
                    sqlite_cur.execute(
                        "DELETE FROM budget_items WHERE budget_id = ? AND description LIKE 'General Fund %'",
                        (sqlite_budget_id,)
                    )
                    for item in data['items']:
                        sqlite_batch_items.append((
                            sqlite_budget_id, item[0], item[1], item[2], 'BARS', item[3]
                        ))
            
            # Execute child items in a single bulk transaction
            if sqlite_batch_items:
                sqlite_cur.executemany(
                    """
                    INSERT INTO budget_items (budget_id, category_type, major_category, amount, account_code, description, embedding)
                    VALUES (?, ?, ?, ?, ?, ?, NULL)
                    """,
                    sqlite_batch_items
                )
                if not pg_conn:
                    saved_item_count = len(sqlite_batch_items)
                    
            sqlite_conn.commit()
            print("  [DB] SQLite write transaction complete.")
        except Exception as e:
            sqlite_conn.rollback()
            print(f"  [ERROR] SQLite transaction failed: {e}")
            
    if pg_conn and pg_cur:
        print("  [DB] Writing to Postgres...")
        try:
            pg_cur.execute("BEGIN")
            
            pg_batch_items = []
            
            for mcag, data in parent_budgets.items():
                std_name = data['std_name']
                etype = data['entity_type']
                tot_rev = data['total_revenue']
                tot_exp = data['total_expenditures']
                beg_bal = data['fund_balance_beginning']
                end_bal = data['fund_balance_ending']
                
                pg_budget_id = None
                
                pg_cur.execute(
                    "SELECT id FROM budgets WHERE jurisdiction_name = %s AND entity_type = %s AND fiscal_year = %s LIMIT 1",
                    (std_name, etype, year)
                )
                row = pg_cur.fetchone()
                if row:
                    pg_budget_id = row[0]
                    pg_cur.execute(
                        """
                        UPDATE budgets 
                        SET total_revenue = %s, total_expenditures = %s, fund_balance_beginning = %s, fund_balance_ending = %s, source_url = %s
                        WHERE id = %s
                        """,
                        (tot_rev, tot_exp, beg_bal, end_bal, source_url, pg_budget_id)
                    )
                else:
                    pg_cur.execute(
                        """
                        INSERT INTO budgets (jurisdiction_name, entity_type, fiscal_year, total_revenue, total_expenditures, fund_balance_beginning, fund_balance_ending, source_url)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                        RETURNING id
                        """,
                        (std_name, etype, year, tot_rev, tot_exp, beg_bal, end_bal, source_url)
                    )
                    pg_budget_id = pg_cur.fetchone()[0]
                    saved_budget_count += 1
                    
                if pg_budget_id and data['items']:
                    pg_cur.execute(
                        "DELETE FROM budget_items WHERE budget_id = %s AND description LIKE 'General Fund %'",
                        (pg_budget_id,)
                    )
                    for item in data['items']:
                        pg_batch_items.append((
                            pg_budget_id, item[0], item[1], item[2], 'BARS', item[3]
                        ))
                        
            if pg_batch_items:
                pg_cur.executemany(
                    """
                    INSERT INTO budget_items (budget_id, category_type, major_category, amount, account_code, description, embedding)
                    VALUES (%s, %s, %s, %s, %s, %s, NULL)
                    """,
                    pg_batch_items
                )
                saved_item_count = len(pg_batch_items)
                
            pg_conn.commit()
            print("  [DB] Postgres write transaction complete.")
        except Exception as e:
            pg_conn.rollback()
            print(f"  [ERROR] Postgres transaction failed: {e}")
            
    if sqlite_conn:
        sqlite_cur.close()
        sqlite_conn.close()
    if pg_conn:
        pg_cur.close()
        pg_conn.close()
        
    print(f"  [PARSER] Finished ingestion. Loaded {saved_budget_count} new budgets and {saved_item_count} breakdowns.")

async def sync_general_fund_budgets(years):
    global SQLITE_PATH
    print(f"🚀 [GENERAL FUND SYNC] Ingesting local General Fund budgets for years: {years}...")
    
    # Copy database locally to prevent Google Drive lock during batch inserts
    temp_db_path = "/Users/thejoshuapenner/.gemini/antigravity/brain/a79b3e23-62f8-4a39-a5f2-fafeec1789e0/scratch/municipal_intent_temp.db"
    import shutil
    original_sqlite_path = SQLITE_PATH
    
    print(f"  [DB COPY] Copying database locally to: {temp_db_path}")
    try:
        shutil.copy2(original_sqlite_path, temp_db_path)
        SQLITE_PATH = temp_db_path
    except Exception as copy_err:
        print(f"  [DB COPY ERROR] Local copy failed: {copy_err}. Proceeding with original path.")
        temp_db_path = None

    for year in years:
        file_path = await download_sao_extract(year)
        if file_path and os.path.exists(file_path):
            try:
                parse_and_ingest_fit_file(file_path, year)
            except Exception as e:
                print(f"  [ERROR] Processing failed for Year {year}: {e}")
            finally:
                try:
                    os.remove(file_path)
                    print(f"  Removed temporary file: {file_path}")
                except Exception:
                    pass
        else:
            print(f"  [ERROR] Download failed for Year {year}. Skipping...")

    # Copy the database back to Google Drive
    if temp_db_path and os.path.exists(temp_db_path):
        print(f"  [DB COPY] Copying updated database back to: {original_sqlite_path}")
        try:
            shutil.copy2(temp_db_path, original_sqlite_path)
            os.remove(temp_db_path)
        except Exception as copy_back_err:
            print(f"  [DB COPY ERROR] Copy back failed: {copy_back_err}")
            
    SQLITE_PATH = original_sqlite_path

if __name__ == "__main__":
    asyncio.run(sync_general_fund_budgets([2022, 2023, 2024, 2025]))
